from openpyxl import load_workbook

wb = load_workbook(filename='C:\\Users\\skolluru\\Documents\\Pharmasourcecom_output_23Nov2018.xlsx',read_only=True)
print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print(ws)
