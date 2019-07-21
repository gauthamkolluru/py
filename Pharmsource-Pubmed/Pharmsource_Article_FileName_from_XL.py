import openpyxl

wb = openpyxl.load_workbook(filename='C:\\Users\\skolluru\\Documents\\Pharmasourcecom_output_23Nov2018.xlsx')
ws = wb[wb.sheetnames[0]]
col = ws['A']
article_file_id = []
for row in col:
    article_file_id.append(row.value)

print(article_file_id)
